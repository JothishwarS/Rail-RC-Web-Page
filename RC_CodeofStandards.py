import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import requests
from io import BytesIO

# ---------- Pages for BS8666:2005 ----------
def bs2005_actual_dia():
    st.markdown("<h3 style='text-align:center;'>Actual Dia - BS8666:2005</h3>", unsafe_allow_html=True)
    st.markdown("<div style='text-align:center;'>"
                "<img src='https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/90ae0691b4786735d50bf779745f209cb70a7703/Actual%20Dia%202005.jpg' width='600'>"
                "</div>", unsafe_allow_html=True)

def bs2005_min_u_bar():
    st.markdown("<h3 style='text-align:center;'>Min U-bar Leg Values - BS8666:2005</h3>", unsafe_allow_html=True)
    raw_pdf = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/90ae0691b4786735d50bf779745f209cb70a7703/BS8666-2005%20MIn%20legs%20of%20U-bars.pdf"
    viewer_url = f"https://docs.google.com/viewer?embedded=true&url={raw_pdf}"
    st.markdown(f"<div style='display:flex; justify-content:center;'>"
                f"<iframe src='{viewer_url}' width='750' height='750'></iframe></div>",
                unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;'><a href='{raw_pdf}' download>📥 Download PDF</a></p>", unsafe_allow_html=True) 

def bs2005_link_sc():
    st.markdown("<h3 style='text-align:center;'>Link SC Min Leg Values - BS8666:2005</h3>", unsafe_allow_html=True)
    excel_url = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/64dcd0c287033258e878d3f44b8c54787610c6cd/Links%20SC%20Min%20values%20for%20BS8666.xlsx"

    shape_ranges_2005 = {
        "Shape Code 33": "A1:D10",
        "Shape Code 51": "A11:E20",
        "Shape Code 63": "A21:E30",
    }

    selected_shape = st.selectbox("Select Shape Code", list(shape_ranges_2005.keys()))

    try:
        response = requests.get(excel_url)
        response.raise_for_status()
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active

        cell_range = shape_ranges_2005[selected_shape]
        data = ws[cell_range]
        table = [[cell.value for cell in row] for row in data]
        max_len = max(len(r) for r in table)
        normalized_table = [r + [None] * (max_len - len(r)) for r in table]

        df = pd.DataFrame(normalized_table[1:], columns=normalized_table[0])

        styled_df = df.style.set_table_styles([
            {'selector': 'th', 'props': [('background-color', '#FFF9C4'), ('text-align', 'center'), ('border', '1px solid black')]},
            {'selector': 'td', 'props': [('text-align', 'center'), ('border', '1px solid gray')]}
        ]).set_properties(**{'padding': '8px'}).hide(axis='index')

        st.markdown("<h4 style='text-align:center;'>Leg Values Table</h4>", unsafe_allow_html=True)

        html_table = styled_df.to_html()
        styled_html = (
        "<div style='display:flex; justify-content:center;'>"
        "<div style='border:1px solid #ccc; padding:10px;'>"
        + html_table +
        "</div></div>"
        )
        st.markdown(styled_html, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"❌ Could not load Excel or shape code range:\n{e}")

def bs2005_shape_codes():
    st.markdown("<h3 style='text-align:center;'>Shape Code List - BS8666:2005</h3>", unsafe_allow_html=True)
    raw_pdf = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/90ae0691b4786735d50bf779745f209cb70a7703/BS8666-2005%20Shape%20Code%20List.pdf"
    viewer_url = f"https://docs.google.com/viewer?embedded=true&url={raw_pdf}"
    st.markdown(f"<div style='display:flex; justify-content:center;'>"
                f"<iframe src='{viewer_url}' width='750' height='750'></iframe></div>",
                unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;'><a href='{raw_pdf}' download>📥 Download PDF</a></p>", unsafe_allow_html=True)

# ---------- Pages for BS8666:2020 ----------
def bs2020_actual_dia():
    st.markdown("<h3 style='text-align:center;'>Actual Dia - BS8666:2020</h3>", unsafe_allow_html=True)
    st.markdown("<div style='text-align:center;'>"
                "<img src='https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/90ae0691b4786735d50bf779745f209cb70a7703/Actual%20Dia%202020.jpg' width='700'>"
                "</div>", unsafe_allow_html=True)

def bs2020_min_u_bar():
    st.markdown("<h3 style='text-align:center;'>Min U-bar Leg Values - BS8666:2020</h3>", unsafe_allow_html=True)
    raw_pdf = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/90ae0691b4786735d50bf779745f209cb70a7703/BS8666-2020%20MIn%20legs%20of%20U-bars.pdf"
    viewer_url = f"https://docs.google.com/viewer?embedded=true&url={raw_pdf}"
    st.markdown(f"<div style='display:flex; justify-content:center;'>"
                f"<iframe src='{viewer_url}' width='750' height='750'></iframe></div>",
                unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;'><a href='{raw_pdf}' download>📥 Download PDF</a></p>", unsafe_allow_html=True)

def bs2020_link_sc():
    st.markdown("<h3 style='text-align:center;'>Link SC Min Leg Values - BS8666:2020</h3>", unsafe_allow_html=True)
    excel_url = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/64dcd0c287033258e878d3f44b8c54787610c6cd/Links%20SC%20Min%20values%20for%20BS8666.xlsx"

    shape_ranges_2020 = {
        "Shape Code 33": "G1:J10",
        "Shape Code 51": "G11:K20",
        "Shape Code 63": "G21:K30",
    }

    selected_shape = st.selectbox("Select Shape Code", list(shape_ranges_2020.keys()))

    try:
        response = requests.get(excel_url)
        response.raise_for_status()
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb.active

        cell_range = shape_ranges_2020[selected_shape]
        data = ws[cell_range]

        # Normalize row length
        table = [[cell.value for cell in row] for row in data]
        max_len = max(len(r) for r in table)
        normalized_table = [r + [None] * (max_len - len(r)) for r in table]

        df = pd.DataFrame(normalized_table[1:], columns=normalized_table[0])

        # Style the DataFrame
        styled_df = df.style.set_table_styles([
            {'selector': 'th', 'props': [('background-color', '#E1F5FE'), ('text-align', 'center'), ('border', '1px solid black')]},
            {'selector': 'td', 'props': [('text-align', 'center'), ('border', '1px solid gray')]}
        ]).set_properties(**{'padding': '8px'}).hide(axis='index')

        st.markdown("<h4 style='text-align:center;'>Leg Values Table</h4>", unsafe_allow_html=True)
        html_table = styled_df.to_html()
        styled_html = (
        "<div style='display:flex; justify-content:center;'>"
        "<div style='border:1px solid #ccc; padding:10px;'>"
        + html_table +
        "</div></div>"
        )
        st.markdown(styled_html, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"❌ Could not load Excel or shape code range:\n{e}")

def bs2020_shape_codes():
    st.markdown("<h3 style='text-align:center;'>Shape Code List - BS8666:2020</h3>", unsafe_allow_html=True)
    raw_pdf = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/90ae0691b4786735d50bf779745f209cb70a7703/BS8666-2020%20Shape%20Code%20List.pdf"
    viewer_url = f"https://docs.google.com/viewer?embedded=true&url={raw_pdf}"
    st.markdown(f"<div style='display:flex; justify-content:center;'>"
                f"<iframe src='{viewer_url}' width='750' height='750'></iframe></div>",
                unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;'><a href='{raw_pdf}' download>📥 Download PDF</a></p>", unsafe_allow_html=True)
# ---------- Pages for Canada ----------
def canada_actual_dia():
    st.markdown("<h3 style='text-align:center;'>Actual Dia - Canada Standards</h3>", unsafe_allow_html=True)
    st.markdown("<div style='text-align:center;'>"
                "<img src='https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/235643b3e9b9e83a33486aa56b831949df8c691e/Canada%20Actual%20bar%20dia.jpg' width='200'>"
                "</div>", unsafe_allow_html=True)

def canada_shape_codes():
    st.markdown("<h3 style='text-align:center;'>Shape Code List - Canada Standards</h3>", unsafe_allow_html=True)
    raw_pdf = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/235643b3e9b9e83a33486aa56b831949df8c691e/Canada%20Standard-Practice-Manual-Standard%20Shape%20code%20List.pdf"
    viewer_url = f"https://docs.google.com/viewer?embedded=true&url={raw_pdf}"
    st.markdown(f"<div style='display:flex; justify-content:center;'>"
                f"<iframe src='{viewer_url}' width='750' height='750'></iframe></div>",
                unsafe_allow_html=True)
    st.markdown(f"<p style='text-align:center;'><a href='{raw_pdf}' download>📥 Download PDF</a></p>", unsafe_allow_html=True)

def canada_lap_values():
    st.markdown("<h3 style='text-align:center;'>Lap & Anchorage Values - Canada Standards</h3>", unsafe_allow_html=True)

    excel_url = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/11041864d188b8bfc934c9393f2636f71d5880aa/Canada%20Lap%20%26%20Anchorage.xlsx"
    options = {
        "Compression Lap": ("Compression", "K21:O36"),
        "Compression Anchorage": ("Compression", "A1:H18"),
        "Tension Lap - Black Bars": ("Tension Black", "A38:L70"),
        "Tension Anchorage - Black Bars": ("Tension Black", "A1:L36"),
        "Tension Lap - Epoxy Bars": ("Tension Epoxy", "A74:L138"),
        "Tension Anchorage - Epoxy Bars": ("Tension Epoxy", "A1:L72"),
    }

    selected_option = st.selectbox("Select Value Type", list(options.keys()))
    sheet_name, cell_range = options[selected_option]

    try:
        response = requests.get(excel_url)
        response.raise_for_status()
        wb = load_workbook(BytesIO(response.content), data_only=True)
        ws = wb[sheet_name]
        data = ws[cell_range]
        table = [[cell.value for cell in row] for row in data]

        blocks = []
        current_paragraph = []
        current_table = []
        state = 'paragraph'

        for row in table:
            non_empty = [str(cell).strip() for cell in row if cell and str(cell).strip().lower() != 'nan']
            if len(non_empty) < 2:
    # Treat as paragraph only if first 2 columns are blank-like
                # still in paragraph
                if state == 'table' and current_table:
                    blocks.append(("table", current_table))
                    current_table = []
                current_paragraph.append(' '.join(non_empty))
                state = 'paragraph'
            else:
                if state == 'paragraph' and current_paragraph:
                    blocks.append(("paragraph", "\n".join(current_paragraph)))
                    current_paragraph = []
                current_table.append(row)
                state = 'table'

        # Final flush
        if current_paragraph:
            blocks.append(("paragraph", "\n".join(current_paragraph)))
        if current_table:
            blocks.append(("table", current_table))

        for block_type, content in blocks:
            if block_type == "paragraph":
                st.markdown(f"<div style='background-color:#FFF3E0; padding:10px; border-left:5px solid #FF9800;'>"
                            f"<strong>{selected_option} Guidelines:</strong><br>{content}</div>", unsafe_allow_html=True)
            elif block_type == "table":
                if not content or len(content) < 2:
                    continue  # Skip if not enough data

                raw_columns = content[0]
                clean_columns = []
                seen = {}
                for col in raw_columns:
                    if col is None:
                        col = "Unnamed"
                    col_str = str(col)
                    count = seen.get(col_str, 0)
                    seen[col_str] = count + 1
                    clean_columns.append(f"{col_str}_{count}" if count else col_str)

                df = pd.DataFrame(content[1:], columns=clean_columns)
                df.replace(["None", "nan", None, pd.NA, float('nan')], "", inplace=True)

                styled_df = df.style.set_table_styles([
                    {'selector': 'th', 'props': [('background-color', '#FFE0B2'), ('text-align', 'center'), ('border', '1px solid black')]},
                    {'selector': 'td', 'props': [('text-align', 'center'), ('border', '1px solid gray')]}
                ]).set_properties(**{'padding': '8px'}).hide(axis='index')

                st.markdown("<h4 style='text-align:center;'>Selected Table</h4>", unsafe_allow_html=True)
                html_table = styled_df.to_html()
                styled_html = (
                    "<div style='display:flex; justify-content:center;'>"
                    "<div style='border:1px solid #ccc; padding:10px;'>"
                    + html_table +
                    "</div></div>"
                )
                st.markdown(styled_html, unsafe_allow_html=True)

    except Exception as e:
        st.error(f"❌ Could not load Excel:\n{e}")
# ---------- Main App ----------
def main():
    st.sidebar.title("Standard Family")
    major_standard = st.sidebar.selectbox("Choose Standard Family:", ["Select...", "BS Code", "Canada Standards"])

    if major_standard == "BS Code":
        st.sidebar.title("BS Code Tools")
        standard = st.sidebar.selectbox("Choose BS Standard:", ["Select...", "BS8666:2005", "BS8666:2020", "Lap Values"])

        if standard == "BS8666:2005":
            st.markdown("<h1 style='text-align:center;'>BS8666:2005 Tools</h1>", unsafe_allow_html=True)
            option = st.sidebar.selectbox("Select tool:", [
                "Actual Dia",
                "Min U-bar Leg Values",
                "Link SC Min Leg Values",
                "Shape Code List"
            ])
            if option == "Actual Dia":
                bs2005_actual_dia()
            elif option == "Min U-bar Leg Values":
                bs2005_min_u_bar()
            elif option == "Link SC Min Leg Values":
                bs2005_link_sc()
            elif option == "Shape Code List":
                bs2005_shape_codes()

        elif standard == "BS8666:2020":
            st.markdown("<h1 style='text-align:center;'>BS8666:2020 Tools</h1>", unsafe_allow_html=True)
            option = st.sidebar.selectbox("Select tool:", [
                "Actual Dia",
                "Min U-bar Leg Values",
                "Link SC Min Leg Values",
                "Shape Code List"
            ])
            if option == "Actual Dia":
                bs2020_actual_dia()
            elif option == "Min U-bar Leg Values":
                bs2020_min_u_bar()
            elif option == "Link SC Min Leg Values":
                bs2020_link_sc()
            elif option == "Shape Code List":
                bs2020_shape_codes()

        elif standard == "Lap Values":
            st.set_page_config(layout="wide")
            st.subheader("EC2 Lap Values by Concrete Grade")

            excel_url = "https://raw.githubusercontent.com/JothishwarS/Rail-RC-Web-Page/main/EC2%20Lap%20Excel.xlsx"

            try:
                response = requests.get(excel_url)
                response.raise_for_status()
                wb = load_workbook(BytesIO(response.content), data_only=True)
                ws = wb.active

                grade_ranges = {
                    "Concrete Grade C20/25": "A1:K9",
                    "Concrete Grade C25/30": "A10:K18",
                    "Concrete Grade C28/35": "A19:K27",
                    "Concrete Grade C30/37": "A28:K36",
                    "Concrete Grade C32/40": "A37:K45",
                    "Concrete Grade C35/45": "A46:K54",
                    "Concrete Grade C40/50": "A55:K63",
                    "Concrete Grade C45/55": "A64:K72",
                    "Concrete Grade C50/60": "A73:K81",
                }

                selected_grade = st.selectbox("Select Concrete Grade", list(grade_ranges.keys()))
                cell_range = grade_ranges[selected_grade]
                data = ws[cell_range]

                table = [[cell.value for cell in row] for row in data]
                df = pd.DataFrame(table[1:], columns=table[0])

                # Style DataFrame
                styled_df = df.style.set_table_styles([
                    {
                        'selector': 'th',
                        'props': [
                            ('background-color', '#C8E6C9'),
                            ('text-align', 'center'),
                            ('border', '1px solid black')
                        ]
                    },
                    {
                        'selector': 'td',
                        'props': [
                            ('text-align', 'center'),
                            ('border', '1px solid gray')
                        ]
                    }
                ]).set_properties(**{'padding': '8px'}).hide(axis='index')

                # Render table
                st.markdown("<h4 style='text-align:center;'>Lap Values Table</h4>", unsafe_allow_html=True)
                html_table = styled_df.to_html()
                styled_html = (
                    "<div style='display:flex; justify-content:center;'>"
                    "<div style='border:1px solid #ccc; padding:10px;'>"
                    + html_table +
                    "</div></div>"
                )
                st.markdown(styled_html, unsafe_allow_html=True)

            except Exception as e:
                st.error(f"❌ Could not load Excel or parse range:\n{e}")

        else:
            st.markdown("Please choose a BS standard from the selector.")

    elif major_standard == "Canada Standards":
        st.sidebar.title("Canada Standards Tools")
        option = st.sidebar.selectbox("Select tool:", [
            "Actual Dia",
            "Min U-bar Leg Values",
            "Shape Code List",
            "Lap & Anchorage Values"
        ])
        if option == "Actual Dia":
            canada_actual_dia()
        elif option == "Min U-bar Leg Values":
            st.info("🔧 Min U-bar Leg Values for Canada Standards will be added shortly.")
        elif option == "Shape Code List":
            canada_shape_codes()
        elif option == "Lap & Anchorage Values":
            canada_lap_values()

    else:
        st.title("Welcome to The Rail Structures RC Code of Standards Web Page")
        st.write("Please choose a code family (BS Code or Canada Standards) to begin.")

if __name__ == "__main__":
    main()
